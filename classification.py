"""첨부 연계표를 읽어 코드 기준으로 정규화한다. 원본 파일에는 쓰지 않는다."""
import pandas as pd
from openpyxl import load_workbook

LEVELS = ['major', 'middle', 'small']


def read_crosswalk(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    taxonomy = pd.DataFrame(
        [row[:6] for row in workbook['1. 상권_업종분류(247)'].iter_rows(min_row=3, values_only=True) if row[4]],
        columns=['major_code', 'major_name', 'middle_code', 'middle_name', 'small_code', 'small_name'])
    rows = []
    for row_number, row in enumerate(workbook['3. 상권_업종연계표(837-247)'].iter_rows(min_row=3, values_only=True), 3):
        if row[8]:
            rows.append([row_number, *row[:9]])
    edges = pd.DataFrame(rows, columns=['sheet_row', 'old_code', 'old_major_name', 'old_middle_name',
        'old_small_name', 'new_code', 'new_major_name', 'new_middle_name', 'new_small_name', 'crosswalk_type'])
    workbook.close()
    assert len(taxonomy) == 247 and taxonomy.small_code.is_unique
    assert set(edges.crosswalk_type) == {'유지', '통합', '분리', '제외', '추가'}
    assert set(edges.new_code.dropna()) == set(taxonomy.small_code)
    # 행 위치만 다른 완전 중복은 연결 관계로 한 번만 사용한다. 원본 행은 별도로 보존한다.
    unique_edges = edges.drop_duplicates(subset=list(edges.columns[1:])).copy()
    return taxonomy, edges, unique_edges


def resolve_candidates(targets, lookup, excluded_possible=False):
    """여러 후보가 공통으로 갖는 수준까지만 확정한다. 후보를 복제·배분하지 않는다."""
    targets = sorted(set(targets))
    result = {'candidate_small_codes': '|'.join(targets), 'candidate_count': len(targets)}
    for level in LEVELS:
        values = {lookup[code][f'{level}_code'] for code in targets}
        result[f'candidate_{level}_codes'] = '|'.join(sorted(values))
        code = next(iter(values)) if len(values) == 1 and not excluded_possible else None
        result[f'normalized_{level}_code'] = code
        result[f'normalized_{level}_name'] = lookup[targets[0]][f'{level}_name'] if code else None
    level = next((level for level in reversed(LEVELS) if result[f'normalized_{level}_code']), None)
    result['mapping_level'] = level or 'none'
    result['mapping_confidence'] = {'small': '소분류확정', 'middle': '중분류확정', 'major': '대분류확정', None: '미확정'}[level]
    return result


def build_mapping(taxonomy, edges):
    lookup = taxonomy.set_index('small_code', drop=False).to_dict('index')
    by_old = {code: group for code, group in edges.dropna(subset=['old_code']).groupby('old_code')}
    assert not set(lookup).intersection(by_old), '신·구 코드가 겹치므로 체계 판별에 추가 근거가 필요합니다.'
    rows = []
    for code in sorted(set(lookup) | set(by_old)):
        if code in lookup:
            # 최신 코드는 구 체계로 역변환하지 않는다. 실제 변환은 항등 매핑이다.
            result = resolve_candidates([code], lookup)
            lineage = sorted(set(edges.loc[edges.new_code.eq(code), 'crosswalk_type']))
            result.update(source_system='247', mapping_type='유지', mapping_basis='최신코드 직접확인',
                          crosswalk_types='|'.join(lineage), is_added_category='추가' in lineage)
        else:
            group = by_old[code]
            targets = sorted(set(group.new_code.dropna()))
            types = set(group.crosswalk_type)
            excluded_possible = '제외' in types and bool(targets)
            result = resolve_candidates(targets, lookup, excluded_possible)
            if not targets and types == {'제외'}:
                mapping_type = '제외'
                result['mapping_confidence'] = '제외확정'
            elif len(targets) > 1 or excluded_possible:
                mapping_type = '분리'
            elif len(targets) == 1:
                mapping_type = '통합' if '통합' in types else '유지'
            else:
                raise ValueError(f'해석할 수 없는 연계 관계: {code}')
            result.update(source_system='837', mapping_type=mapping_type, mapping_basis='837→247 연계',
                          crosswalk_types='|'.join(sorted(types)), is_added_category=False)
        result['source_small_code'] = code
        rows.append(result)
    mapping = pd.DataFrame(rows)
    assert mapping.source_small_code.is_unique
    return mapping


def normalize_records(frame, mapping):
    result = frame.merge(mapping, on='source_small_code', how='left', validate='many_to_one')
    unknown = result.mapping_type.isna()
    result.loc[unknown, ['source_system', 'mapping_type', 'mapping_confidence', 'mapping_level', 'mapping_basis']] = [
        '미확인', '미매핑', '미확정', 'none', '연계표에 없는 코드']
    return result


def aggregate_level(records, taxonomy, edges, level, group_columns):
    """확정 집계와 하한을 구분한다. 구 체계의 추가 업종·불확정 분리 후보를 0으로 채우지 않는다."""
    normalized_code = f'normalized_{level}_code'
    target_column = f'{level}_code'
    targets = taxonomy[[target_column, f'{level}_name']].drop_duplicates().set_index(target_column)
    additions = set(edges.loc[edges.crosswalk_type.eq('추가'), 'new_code'])
    addition_groups = set(taxonomy.loc[taxonomy.small_code.isin(additions), target_column])
    rows = []
    for keys, group in records.groupby(group_columns, sort=True):
        keys = keys if isinstance(keys, tuple) else (keys,)
        counts = group[normalized_code].dropna().value_counts()
        uncertain = group.loc[group[normalized_code].isna() & ~group.mapping_type.eq('제외')]
        affected = set()
        for value in uncertain[f'candidate_{level}_codes'].dropna():
            affected.update(value.split('|') if value else [])
        unknown = group.mapping_type.eq('미매핑').any()
        old_present = group.source_system.eq('837').any()
        for code in targets.index:
            reasons = []
            if unknown:
                reasons.append('미매핑코드 존재')
            if code in affected:
                reasons.append('분리후보 미확정')
            if old_present and code in addition_groups:
                reasons.append('추가범위의 과거대응 없음')
            count = int(counts.get(code, 0))
            rows.append(dict(zip(group_columns, keys)) | {'category_code': code,
                'category': targets.loc[code, f'{level}_name'], 'store_count': None if reasons else count,
                'confirmed_count': count, 'coverage_status': '|'.join(reasons) if reasons else '관측체계 내 확정집계'})
    return pd.DataFrame(rows)
